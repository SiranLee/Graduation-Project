from django.shortcuts import render, HttpResponse
from django.http import FileResponse
import os
from django.conf import settings
from django.utils.encoding import escape_uri_path
import cv2
from .models import Teacher, Student, Department, ImageStore, Course, Source, File, StudentStore, Mclass, StudentImage, TeacherImage, Admin, AdminImage, TeacherStore, StudentTime, StudentTimeNow, StudentWork, StudentRoutesMap, TeacherRoutesMap, Routes, StagingFile, StagingConvertPDF
import json
import base64

# Create your views here.
def upfile(request):
    return render(request, 'teaching_system/upfile.html')

def savefile(request):
    if request.method == 'POST':
        # 获取资源类型（获取value值）
        resource_type = request.POST.get('resource_type')
        print(resource_type)
        #合成文件夹路径
        if resource_type is not None:
            fileDir = os.path.join(settings.MEDIA_ROOT, resource_type)
            # 取文件
            f = request.FILES['file']
            # 合成文件在服务器端的路径
            filePath = os.path.join(fileDir, f.name)
            with open(filePath, 'wb') as fp:
                for info in f.chunks():
                    # f.chunks()文件分块接收
                    fp.write(info)
            return HttpResponse('上传成功')
        else:
            return HttpResponse('请选择上传类型')
    else:
        return HttpResponse('上传失败')


def download(request):
    file = open('static/upfile/image/游戏安装教程.png', 'rb')
    the_file_name = file.name
    print(the_file_name)
    response = FileResponse(file)
    response['Content-Type'] = 'application/octet-stream'
    response['Content-Disposition'] = 'attachment;filename*=utf-8''{}'.format(escape_uri_path(the_file_name))
    return response


def video(request):
    # 0表示使用电脑摄像头
    cap = cv2.VideoCapture(0)
    fourcc = cv2.VideoWriter_fourcc(*'XVID')
    out = cv2.VideoWriter('output.avi', fourcc, 20.0, (640, 480))

    while (cap.isOpened()):
        ret, frame = cap.read()
        if ret == True:
            # 视频翻转
            # 1	水平翻转
            # 0	垂直翻转
            # -1	水平垂直翻转
            frame = cv2.flip(frame, 0)

            out.write(frame)
            cv2.imshow('frame', frame)
            if cv2.waitKey(1) & 0xFF == ord('q'):
                break
        else:
            break
    cap.release()
    out.release()
    cv2.destroyAllWindows()


import hashlib
def login(request):
    username = eval(request.body)['username']
    print(username)
    password = eval(request.body)['password']
    # print(password)
    obj = hashlib.md5(username.encode())  # 实例化md5的时候可以给传个参数，这叫加盐
    obj.update(password.encode("utf-8"))  # 是再加密的时候传入自己的一块字节，
    secret = obj.hexdigest()
    if username is not None:
        # sql查询
        usernameInfo = Teacher.teacherManage.filter(tno=username, isDelete=False)
        if len(usernameInfo) > 0:
            #有这个用户
            usernameInfo = usernameInfo[0]
            # print(secret)
            # print(usernameInfo.password)
            if usernameInfo.password == secret:
                # token = str(base64.b64encode(username.encode()))
                token = base64.b64encode(username.encode('utf-8')).decode()
                request.session['role'] = 'teacher'
                # print(request.session.get('role'))
                request.session['id'] = username
                # print(request.session.get('id'))
                request.session['token'] = token
                # print(request.session.get('token'))
                response = HttpResponse(json.dumps({'code': 20000,
                                                    'data': {'token':token,'message':'验证成功','role':['teacher']}
                                                    }))
            else:
                response = HttpResponse(
                    json.dumps({'code': 0, 'data': {'token': "",'message':'密码错误'}}))
        else:
            studentusers = Student.studentManager.filter(sno=username)
            if len(studentusers) > 0:
                #有这个用户
                studentuser = studentusers[0]
                if studentuser.password == secret:
                    token = base64.b64encode(username.encode('utf-8')).decode()
                    request.session['role'] = 'student'
                    request.session['id'] = username
                    request.session['token'] = token
                    # 学生登陆时进行自己写入自己的路由
                    stuRoutes = StudentRoutesMap.studentRoutesMapManager.filter(stu_no=username)
                    if len(stuRoutes) == 0:
                        for index in range(1, 10):
                            studentRouteMap = StudentRoutesMap.studentRoutesMapManager.createStudentRoutesMap(username, index)
                            studentRouteMap.save()
                    
                    response = HttpResponse(json.dumps({'code': 20000,
                                                        'data': {'token':token,'message':'验证成功', 'role':['student']}
                                                        }))
                else:
                    response = HttpResponse(
                        json.dumps({'code': 0, 'data': {'token': "",'message':'密码错误'}}))
            else:
                admin = Admin.adminManage.filter(ano=username)
                # print(admin)
                if len(admin) > 0:
                    #超级管理员
                    admin = admin[0]
                    print(admin.aname)
                    if admin.password == secret:
                        token = base64.b64encode(username.encode('utf-8')).decode()
                        request.session['role'] = 'admin'
                        request.session['id'] = username
                        request.session['token'] = token
                        response = HttpResponse(json.dumps({'code': 20000,
                                                            'data': {'token': token, 'message': '验证成功', 'role':['admin']}
                                                            }))
                    else:
                        response = HttpResponse(
                    json.dumps({'code': 0, 'data': {'token':'','message': '密码错误'}}))
                else:
                    response = HttpResponse(
                        json.dumps({'code': 0, 'data': {'token': '', 'message': '当前用户不存在'}}))
    else:
        response = HttpResponse(
            json.dumps({'code': 0, 'data': {'token':'','message': '用户名为空'}}))

    response['Access - Control - Allow - Origin'] = "*"
    response["Access - Control - Allow - Methods"] = "POST, GET, OPTIONS"
    response["Access - Control - Max - Age"] = "1000"
    response["Access - Control - Allow - Headers"] = "*"
    response["X-Frame-Options"] = "allow-from *"
    print(response)
    return response

def logout(request):
    token = eval(request.body)['token']
    # id = int(base64.b64decode(token.encode()).decode())
    id = base64.b64decode(token.encode('utf-8')).decode()
    user = Teacher.teacherManage.filter(tno=id)
    # print(user)
    if len(user) > 0:
        # 有这个用户
        pass
    else:
        studentuser = Student.studentManager.filter(sno=id)
        if len(studentuser) > 0:
            # 有这个用户
            time = datetime.datetime.now()
            # print(time)
            # print(type(time))

            # 查询时间
            current_date = str(datetime.datetime.now()).split()[0]
            stutimenow = StudentTimeNow.studentTimeNowManager.filter(sno=id)[0]
            stutime = StudentTime.studentTimeManager.filter(sno=studentuser[0].sno)
            label = False
            for item in stutime:
                if item.time == current_date:
                    item.timelong = item.timelong + float((time - stutimenow.time).seconds)
                    stutime.save()
                    label = True
                    stutimenow.delete()
            if label is False:
                stutime = StudentTime.studentTimeManager.createStudentTime(studentuser[0].sno, float((time - stutimenow.time).seconds), current_date)
                stutime.save()
                stutimenow.delete()
        else:
            admin = Admin.adminManage.filter(ano=id)
            if len(admin) > 0:
                pass
    return HttpResponse(json.dumps({
        'code': 20000,
        'data': {
            'message': "ok"
        }
    }))

def test(request):
    token = request.POST.get('token')
    # id = int(base64.b64decode(token.encode()).decode())
    id = base64.b64decode(token.encode()).decode()
    user = Student.studentManager.filter(sno=id)
    if len(user) > 0:
        time = datetime.datetime.now()
        # 查询时间
        current_date = str(datetime.datetime.now()).split()[0]
        stutimenow = StudentTimeNow.studentTimeNowManager.filter(sno=id)[0]
        stutime = StudentTime.studentTimeManager.filter(sno=user[0].sno)
        label = False
        for item in stutime:
            if item.time == current_date:
                item.timelong = item.timelong + float((time - stutimenow.time).seconds)
                stutime.save()
                label = True
                stutimenow.delete()
        if label is False:
            stutime = StudentTime.studentTimeManager.createStudentTime(user[0].sno,
                                                                       float((time - stutimenow.time).seconds),
                                                                       current_date)
            stutime.save()
            stutimenow.delete()
        # #查询时间
        # current_date = str(datetime.datetime.now()).split()[0]
        # stutimenow = StudentTimeNow.studentTimeNowManager.filter(sno=id)[0]
        # stutime = StudentTime.studentTimeManager.createStudentTime(user[0].sno, float((time-stutimenow.time).seconds), current_date)
        # stutime.save()
        # stutimenow.delete()
    return HttpResponse(json.dumps({
        'code': 20000,
        'data': {
            'message': "ok"
        }
    }))

import datetime
def check(request):
    token = request.GET.get('token')
    # id = int(base64.b64decode(token.encode()).decode())
    id = base64.b64decode(token.encode()).decode()
    user = Teacher.teacherManage.filter(tno=id)
    if len(user) > 0:
        #有这个用户
        return HttpResponse(json.dumps({'code': 20000, 'data': {
            'avatar': user[0].image,
            'name': user[0].tname,
            'id': user[0].tno,
            'roles': ['teacher'],
            'major': user[0].department.dname
        }}))
    else:
        studentusers = Student.studentManager.filter(sno=id)
        if len(studentusers) > 0:
            #有这个用户
            #保存时间
            # current_date = str(datetime.datetime.now()).split()[0]
            savetime = StudentTimeNow.studentTimeNowManager.createStudentTimeNow(studentusers[0].sno)
            savetime.save()
            # 学生登录时查询已经选择的课程
            course_nos = []
            for student in studentusers:
                mclass = student.mclass
                course_nos.append(mclass.con.no)
            return HttpResponse(json.dumps({'code': 20000, 'data': {
                        'avatar': studentusers[0].image,
                        'name': studentusers[0].sname,
                        'id': studentusers[0].sno,
                        'roles': ['student'],
                        'major': studentusers[0].department.dname,
                        'course_nos': course_nos
                    }}))
        else:
            admin = Admin.adminManage.filter(ano=id)
            if len(admin) > 0:
                return HttpResponse(json.dumps({'code': 20000, 'data': {
                    'avatar': admin[0].image,
                    'name': admin[0].aname,
                    'id': admin[0].ano,
                    'roles': ['admin'],
                }}))

def teacher_getcourse(request):
    token = request.GET.get('teacher_id')
    # print(token)
    # id = int(base64.b64decode(token.encode()).decode())
    teacher = Teacher.teacherManage.filter(tno=token)[0]
    courses = teacher.course_set.all()
    list = []
    for item in courses:
        dic = {
            'course_id': item.no,
            'cover_link': item.image,
            'title': item.name
        }
        list.append(dic)
    return HttpResponse(json.dumps({
        'code':20000,
        'data':{
            'course':list
        }
    }))

def get_course(request):
    course_id = request.GET.get('course_id')
    course = Course.courseManager.filter(no=course_id)[0]
    return HttpResponse(json.dumps({
        'code': 20000,
        'data': {
            'cover': course.image,
            'intro': course.intro,
            'teacher': course.tno.tname,
            'title': course.name
        }
    }))

def get_sourcetype(request):
    types = list(Source.sourceManager.all())
    type = []
    taskTypes = []
    for item in types:
        if types.index(item) > 2:
            dic1 = {
                'types_value': item.sno,
                'types_label': item.sname,
            }
            taskTypes.append(dic1)
        else:
            dic = {
                'label': item.sname,
                'value': item.sno
            }
            type.append(dic)
    return HttpResponse(json.dumps({
        'code': 20000,
        'data': {
            'types':type,
            'taskTypes': taskTypes
        }
    }))

#上传资源
def up_source(request):
    course_id = request.POST.get('course_id')
    # print(course_id)
    course = Course.courseManager.filter(no=course_id)[0]
    department = course.dno
    # print(course)
    teacher = course.tno
    # print(teacher)
    describe = request.POST.get('describe')
    source_type = request.POST.get('source_type')
    not_available_2_all = request.POST.get('not_available2all')
    source = Source.sourceManager.get(sno=source_type)
    # print(source)
    count = request.POST.get('count')
    fileList = []
    for i in range(int(count)):
        file1 = request.FILES.get('file' + str(i))
        fileDir = os.path.join(settings.MEDIA_ROOT, 'stagingSource')
        filePath = os.path.join(fileDir, file1.name)
        with open(filePath, 'wb') as fp:
            for info in file1.chunks():
                fp.write(info)
        # print(os.path.join('/upfile/source', file1.name))
        #f = StagingFile.stagingFileManager.createStagingFile(r'/upfile/stagingSource/'+file1.name, request.POST.get('title'), file1.name, describe, 1, not_available_2_all, course, teacher, source, department)
        f = StagingFile.stagingFileManager.createStagingFile(r'/upfile/stagingSource/'+file1.name, request.POST.get('title'), file1.name, describe, 1, not_available_2_all, ' ', course, teacher, source, department)
        # f = File.fileManager.createFile(describe, r'/upfile/source/' + file1.name, request.POST.get('title'), file1.name, 0, department.dno, not_available_2_all, teacher, source, course)
        f.save()
        fileList.append(f)

    listItem = []
    count = len(fileList)
    for item in fileList:
        dic = {
            'upload_date': str(item.up_time).split()[0],
            'upload_type': item.sno.sname,
            'upload_title': item.title,
            'upload_intro': item.fileDes,
            'upload_filename': item.filename,
            'upload_filelink': item.url,
            'upload_id': item.id
        }
        listItem.append(dic)
    types = list(Source.sourceManager.all())
    type = []
    taskTypes = []
    for item in types:
        if types.index(item) > 2:
            dic1 = {
                'types_value': item.sno,
                'types_label': item.sname,
            }
            taskTypes.append(dic1)
        else:
            dic = {
                'label': item.sname,
                'value': item.sno
            }
            type.append(dic)
    return HttpResponse(json.dumps({
        'code': 20000,
        'data': {
            'listTotal': count,
            'list': listItem,
            'types': type
        }
    }))

def get_source(request):
    course_id = request.GET.get('course_id')
    # print(course_id)
    current_page = int(request.GET.get('page'))
    # print(page)
    page_size = int(request.GET.get('limit'))
    # print(limit)
    course = Course.courseManager.filter(no=course_id)[0]
    # print(course)
    sources = course.file_set.all()
    # print(sources)
    totalCount = len(sources)
    result = []
    # if int(page)*int(limit) < slen:
    #     sources = sources[(int(page) - 1) * limit:(int(page) - 1) * limit + limit]
    # if int(page)*int(limit) - slen < limit and int(page)*int(limit) > slen:
    #     sources = sources[(int(page)-1)*limit:slen - (int(page)-2)*limit]
    # if int(page)*int(limit) - slen > limit:
    #     return HttpResponse(json.dumps({
    #         'code': 20000,
    #         'data': {
    #             'listTotal': 0,
    #             'list': [],
    #         }
    #     }))
    
    # count = len(sources)
    split_page(sources, current_page, page_size, totalCount, result)

    return HttpResponse(json.dumps({
        'code': 20000,
        'data': {
            'listTotal':totalCount,
            'list': result,
        }
    }))

def split_page(sources, current_page, page_size, totalCount, result):
    start_index = (current_page - 1) * page_size + 1
    if start_index - 1 + page_size >= totalCount:
        for source in sources[start_index-1:totalCount]:
            # 此时不足一页或者刚好一页
            result.append(create_source_dic(source))
    else:
        #此时超出一页,只获取一页的数据
        for source in sources[start_index-1:start_index+page_size-1]:
            result.append(create_source_dic(source))
            
def create_source_dic(item):
    dic = {
        'upload_date': str(item.time).split()[0],
        'upload_type': item.sno.sname,
        'upload_title': item.title,
        'upload_intro': item.describe,
        'upload_filename': item.filename,
        'upload_filelink': item.url,
        'upload_id': item.id,
        'read_limit': item.not_available2all
    }
    return dic



def del_source(request):
    source_id = request.GET.get('source_id')
    # print(source_id)
    file = File.fileManager.filter(pk=int(source_id))[0]
    # print(file)
    # file.isDelete = 1
    # file.save()
    fileDir = os.path.join(settings.MEDIA_ROOT, 'source')
    studentstorepath = os.path.join(fileDir, file.url.split('/')[-1])
    os.remove(studentstorepath)
    file.delete()
    return HttpResponse(json.dumps({
        'code': 20000,
        'data': {
            'message': 'ok',
        }
    }))

def update_info(request):
    title = request.POST.get('title')
    teacher_id = request.POST.get('teacher_id')
    intro = request.POST.get('intro')
    # cover = request.POST.get('cover')
    course_id = request.POST.get('course_id')
    file_content = ContentFile(request.FILES['cover'].read())
    
    img = ImageStore(name=request.FILES['cover'].name, img=request.FILES['cover'])
    img.save()
    imagepath = r'/upfile/course_img/' + request.FILES['cover'].name
    course = Course.courseManager.filter(no=course_id)[0]
    course.title = title
    course.tno = Teacher.teacherManage.get(tno=teacher_id)
    course.intro = intro
    course.image = imagepath
    course.save()
    response = HttpResponse(json.dumps({
        'code': 20000,
        'data': {
            'filePath': imagepath,
            'info': {
                'intro': intro,
                'teacher': course.tno.tname,
                'title': title
            }
        }
    }))
    response['Access - Control - Allow - Origin'] = "*"
    response["Access - Control - Allow - Methods"] = "POST, GET, OPTIONS"
    response["Access - Control - Max - Age"] = "1000"
    response["Access - Control - Allow - Headers"] = "*"
    response["X-Frame-Options"] = "allow-from *"
    return response

def get_department(request):
    id = request.GET.get('id')
    teacher = Teacher.teacherManage.get(tno=id, isDelete=False)
    proper_majors = Course.courseManager.filter(tno=teacher).values('dno').distinct()

    majors = []
    for item in proper_majors:
        major = Department.departmentManage.get(dno=item['dno'])
        dic = {
            'title': major.dname,
            'major_id': major.dno
        }
        majors.append(dic)
    return HttpResponse(json.dumps({
        'code': 20000,
        'data': {
            'majors': majors,
            'link': r'/upfile/template/students.xlsx'
        }
    }))

def get_all_majors(request):
    all_department = Department.departmentManage.all()
    result = []
    for department in all_department:
        dic = {
            'title': department.dname,
            'major_id': department.dno
        }
        result.append(dic)
    
    return HttpResponse(json.dumps({
        'code': 20000,
        'data': {
            'majors': result
        }
    }))

def get_dallcourse(request):
    major_id = request.GET.get("major_id")
    page = int(request.GET.get('page'))
    limit = request.GET.get('limit')

    department = Department.departmentManage.get(dno=major_id)
    courses = department.course_set.all()
    clen = len(courses)
    if int(page) * int(limit) < clen:
        courses = courses[(int(page) - 1) * 7:(int(page) - 1) * 7 + 7]
    if int(page) * int(limit) - clen < 7 and int(page) * int(limit) > clen:
        courses = courses[(int(page) - 1) * 7:clen - (int(page) - 2) * 7]
    if int(page) * int(limit) - clen > 7:
        return HttpResponse(json.dumps({
            'code': 20000,
            'data': {
                'total': 0,
                'courses': [],
            }
        }))

    list = []
    count = len(courses)
    for item in courses:
        if item.isDelete:
            continue
        course_count = len(item.file_set.all())
        dic = {
            'course_cover': item.image,
            'course_name': item.name,
            'course_teacher': item.tno.tname,
            'course_id': item.no,
            'upload_count': course_count,
        }
        list.append(dic)
    return HttpResponse(json.dumps({
        'code': 20000,
        'data': {
            'total': clen,
            'courses': list,
        }
    }))

def course_info(request):
    course_id = request.GET.get('course_id')
    # print(token)
    # id = int(base64.b64decode(token.encode()).decode())
    course = Course.courseManager.filter(no=course_id)[0]
    myclass = course.mclass_set.all()
    count = 0
    for item in myclass:
        students = item.student_set.all()
        count = count + len(students)
    return HttpResponse(json.dumps({
        'code':20000,
        'data':{
            'source_total':len(course.file_set.all()),
            'stu_count': count,
            'course_name': course.name,
            'course_teacher': course.tno.tname,
            'course_intro': course.intro,
            'course_cover': course.image
        }
    }))

def get_sourcebytype(request):
    course_id = request.GET.get('course_id')
    # print(course_id)
    source_type = request.GET.get('source_type')
    # print(source_type)
    page = request.GET.get('page')
    # print(page)
    limit = request.GET.get('limit')
    # print(limit)
    course = Course.courseManager.filter(no=course_id)[0]
    # print(course)
    sources = course.file_set.all()
    files = []
    for item in sources:
        if item.sno.sno == source_type:
            files.append(item)
    # print(sources)
    slen = len(files)
    if int(page)*int(limit) < slen:
        files = files[(int(page) - 1) * 5:(int(page) - 1) * 5 + 5]
    if int(page)*int(limit) - slen < 5 and int(page)*int(limit) > slen:
        files = files[(int(page)-1)*5:slen - (int(page)-2)*5]
    if int(page)*int(limit) - slen > 5:
        return HttpResponse(json.dumps({
            'code': 20000,
            'data': {
                'listTotal': 0,
                'list': [],
            }
        }))
    list = []
    count = len(files)
    for item in files:
        dic = {
            'upload_id': item.id,
            'upload_date': str(item.time).split()[0],
            'upload_type': item.sno.sname,
            'upload_title': item.title,
            'upload_intro': item.describe,
            'upload_filename': item.filename,
            'upload_filelink': item.url,

        }
        list.append(dic)
    return HttpResponse(json.dumps({
        'code': 20000,
        'data': {
            'total':slen,
            'course_resourceList': list,
        }
    }))

def tsearch_source(request):
    key_word = request.GET.get('key_word')
    course_id = request.GET.get('course_id')
    page = request.GET.get('page')
    limit = request.GET.get('limit')

    files = File.fileManager.all()
    source = []
    for item in files:
        if key_word in item.title:
            source.append(item)
    slen = len(source)
    if int(page) * int(limit) < slen:
        source = source[(int(page) - 1) * 5:(int(page) - 1) * 5 + 5]
    if int(page) * int(limit) - slen < 5 and int(page) * int(limit) > slen:
        source = source[(int(page) - 1) * 5:slen - (int(page) - 2) * 5]
    if int(page) * int(limit) - slen > 5:
        return HttpResponse(json.dumps({
            'code': 20000,
            'data': {
                'listTotal': 0,
                'list': [],
            }
        }))
    list = []
    count = len(source)
    for item in source:
        dic = {
            'upload_id': item.id,
            'upload_date': str(item.time).split()[0],
            'upload_type': item.sno.sname,
            'upload_title': item.title,
            'upload_intro': item.describe,
            'upload_filename': item.filename,
            'upload_filelink': item.url,

        }
        list.append(dic)
    return HttpResponse(json.dumps({
        'code': 20000,
        'data': {
            'total': slen,
            'course_resourceList': list,
        }
    }))

def get_courseinfo(request):
    major_id = request.GET.get('major_id')
    # print(major_id)
    department = Department.departmentManage.get(dno=major_id)
    courses = department.course_set.all()
    list = []
    for item in courses:
        dic = {
            'title': item.name,
            'course_id': str(item.pk)
        }
        list.append(dic)
    return HttpResponse(json.dumps({
        'code': 20000,
        'data': {
            'courses': list
        }
    }))


from openpyxl.reader.excel import load_workbook
def load_studentinfo(request):
    major_id = request.POST.get('major_id')
    department = Department.departmentManage.get(dno=major_id)
    course_id = request.POST.get('course_id')
    course = Course.courseManager.get(no=course_id)
    file_content = ContentFile(request.FILES['excel'].read())
    studentStore = StudentStore(name=request.FILES['excel'].name, src=request.FILES['excel'])
    studentStore.save()
    fileDir = os.path.join(settings.MEDIA_ROOT, 'studentStore')
    studentstorepath = os.path.join(fileDir, request.FILES['excel'].name)
    # print(studentstorepath)
    # 打开文件
    file = load_workbook(filename=studentstorepath)
    # 所有表格的名称
    # sheets = file.sheetnames

    # 拿出一个表格
    sheet = file.worksheets[0]

    # 读取数据
    sumlist = []
    diclist = []
    mclass = None
    for lineNum in range(1, sheet.max_row + 1):
        # print(lineNum)
        lineList = []
        for columnNum in range(1, sheet.max_column + 1):
            # 拿数据
            value = sheet.cell(row=lineNum, column=columnNum).value
            lineList.append(value)
        if lineNum == 2:
            #创建一个班级
            # department = Department.departmentManage.filter(dname=lineList[2])
            # print(lineList)
            #先获取班级
            mclass = Mclass.mclassManager.filter(cno='200' + course_id)
            if len(mclass) > 0:
                mclass = mclass[0]
            else:
                mclass = Mclass.mclassManager.createMclass('200' + course_id, course.name, str(lineList[1]), department, course)
                mclass.save()
        if lineNum > 1:
            dic = {
                'stu_number': lineList[0],
                'stu_grade': lineList[1],
                'stu_major': lineList[2],
                'stu_class': lineList[3],
                'name': lineList[4],
                'course': course.name
            }
            diclist.append(dic)
            try:
                
                student = Student.studentManager.createStudent(lineList[0], lineList[1], lineList[4], '/upfile/template/icon.jpg', str(lineList[3]), "", mclass, department)
                # print('here')
                student.save()
            except Exception as e:
                # print(e)
                pass
            sumlist.append(lineList)

    os.remove(studentstorepath)
    # print(mclass)
    students = Student.studentManager.filter(isDelete = False, mclass=mclass)
    studic = []
    for item in students:
        dic = {
            'stu_number': item.sno,
            'stu_grade': item.grade,
            'stu_major': department.dname,
            'stu_class': item.cs,
            'name': item.sname,
            'course': course.name
        }
        studic.append(dic)
        
    return HttpResponse(json.dumps({
        'code': 20000,
        'data': {
            'total': len(students),
            'stus': studic
        }
    }))

# 学生上传后写入学生路由和学生班级
def create_class_set_routes(request):
    course_id = json.loads(request.body)['course_id']
    stu_grade = json.loads(request.body)['grade']
    item_snos = json.loads(request.body)['snos']
    course = Course.courseManager.get(isDelete=False, no=course_id)
    mclass = Mclass.mclassManager.get(isDelete=False, cno='200'+course_id)
        # mclass = Mclass.mclassManager.createMclass("200" + course_id, course.name, stu_grade, course.dno, course)
        # mclass.save()
        # mclass = Mclass.mclassManager.get(cno="200"+course_id)
        #向学生表中写密码和班级
    for sno in item_snos:
        student = Student.studentManager.get(isDelete=False, sno=sno, mclass=mclass)
        obj = hashlib.md5(sno.encode())  # 实例化md5的时候可以给传个参数，这叫加盐
        obj.update("123456".encode("utf-8"))  # 是再加密的时候传入自己的一块字节，
        secret = obj.hexdigest()
        student.password = secret
        student.mclass = mclass
        student.save()
    
    return HttpResponse(json.dumps({
        'code': 20000,
        'data': {
            'message': 'ok'
        }
    }))

def get_student(request):
    major_id = request.GET.get('major_id')
    dp = Department.departmentManage.get(dno=major_id)
    course_id = request.GET.get('course_id')
    cs = Course.courseManager.get(no=course_id)
    page = int(request.GET.get('page'))
    limit = int(request.GET.get('limit'))

    students = Student.studentManager.all()
    # print(students)

    source = []
    for item in students:
        if item.department.dno == major_id and item.mclass.con.no == course_id:
            # print('1')
            source.append(item)
    slen = len(source)
    # if page*limit < slen :
    #     source = source[(int(page) - 1) * limit:(int(page) - 1) * limit + limit]
    if int(page) * int(limit) < slen:
        source = source[(int(page) - 1) * limit:(int(page) - 1) * limit + limit]
    if int(page) * int(limit) - slen < limit and int(page) * int(limit) > slen:
        # print(1)
        source = source[(int(page) - 1) * limit:slen]
        # print(source)
    if int(page) * int(limit) - slen > limit:
        return HttpResponse(json.dumps({
            'code': 20000,
            'data': {
                'total': 0,
                'stus': [],
            }
        }))

    # print(source)
    list = []
    count = len(source)
    for item in source:
        dic = {
            'stu_number': item.sno,
            'stu_grade': item.grade,
            'stu_major': dp.dname,
            'stu_class': item.cs,
            'name': item.sname,
            'course': cs.name
        }
        list.append(dic)
    # print(list)
    return HttpResponse(json.dumps({
        'code': 20000,
        'data': {
            'total': slen,
            'stus': list,
        }
    }))

def search_student(request):
    major_id = request.GET.get('major_id')
    dp = Department.departmentManage.get(dno=major_id)
    course_id = request.GET.get('course_id')
    cs = Course.courseManager.get(no=course_id)
    keyword = request.GET.get('keyword')
    page = request.GET.get('page')
    limit = int(request.GET.get('limit'))

    students = Student.studentManager.filter(department=dp)

    source = []
    for item in students:
        if item.sname == keyword and item.mclass.con.no == course_id:
            source.append(item)
    slen = len(source)
    if int(page) * int(limit) < slen:
        source = source[(int(page) - 1) * limit:(int(page) - 1) * limit + limit]
    if int(page) * int(limit) - slen < limit and int(page) * int(limit) > slen:
        source = source[(int(page) - 1) * limit:slen - (int(page) - 2) * limit]
    if int(page) * int(limit) - slen > limit:
        return HttpResponse(json.dumps({
            'code': 20000,
            'data': {
                'total': 0,
                'stus': [],
            }
        }))

    list = []
    count = len(source)
    for item in source:
        dic = {
            'stu_number': item.sno,
            'stu_grade': item.grade,
            'stu_major': dp.dname,
            'stu_class': item.cs,
            'name': item.sname,
            'course': cs.name
        }
        list.append(dic)
    return HttpResponse(json.dumps({
        'code': 20000,
        'data': {
            'total': slen,
            'stus': list,
        }
    }))

def search_with_no(request):
    major_id = request.GET.get('major_id')
    dp = Department.departmentManage.get(dno=major_id)
    course_id = request.GET.get('course_id')
    cs = Course.courseManager.get(no=course_id)
    stu_no = request.GET.get('stuNo')
    page = request.GET.get('page')
    limit = int(request.GET.get('limit'))

    students = Student.studentManager.filter(department=dp)

    source = []
    for item in students:
        if item.sno == stu_no and item.mclass.con.no == course_id:
            source.append(item)
    slen = len(source)
    if int(page) * int(limit) < slen:
        source = source[(int(page) - 1) * limit:(int(page) - 1) * limit + limit]
    if int(page) * int(limit) - slen < limit and int(page) * int(limit) > slen:
        source = source[(int(page) - 1) * limit:slen - (int(page) - 2) * limit]
    if int(page) * int(limit) - slen > limit:
        return HttpResponse(json.dumps({
            'code': 20000,
            'data': {
                'total': 0,
                'stus': [],
            }
        }))

    list = []
    count = len(source)
    for item in source:
        dic = {
            'stu_number': item.sno,
            'stu_grade': item.grade,
            'stu_major': dp.dname,
            'stu_class': item.cs,
            'name': item.sname,
            'course': cs.name
        }
        list.append(dic)
    return HttpResponse(json.dumps({
        'code': 20000,
        'data': {
            'total': slen,
            'stus': list,
        }
    }))
def reset_password(request):
    # print(request.body)
    # student_id = request.POST.get('id')
    student_id = eval(request.body)['id']
    # print(student_id)
    # pwd = request.POST.get('pwd')
    pwd = eval(request.body)['pwd']

    obj = hashlib.md5(student_id.encode())  # 实例化md5的时候可以给传个参数，这叫加盐
    obj.update(pwd.encode("utf-8"))  # 是再加密的时候传入自己的一块字节，
    secret = obj.hexdigest()


    students = Student.studentManager.filter(sno=student_id)
    for student in students:
        student.password = secret
        student.save()

    return HttpResponse(json.dumps({
        'code': 20000,
        'data': {
            'message': 'ok'
        }
    }))

def reset_passwordall(request):
    id = eval(request.body)['id']
    pwd = eval(request.body)['pwd']
    role = eval(request.body)['role']

    obj = hashlib.md5(id.encode())  # 实例化md5的时候可以给传个参数，这叫加盐
    obj.update(pwd.encode("utf-8"))  # 是再加密的时候传入自己的一块字节，
    npwd = obj.hexdigest()
    # print(npwd)

    if role == 's':
        student = Student.studentManager.get(sno=id)
        student.password = npwd
        student.save()
    if role == 't':
        teacher = Teacher.teacherManage.get(tno=id)
        teacher.password = npwd
        teacher.save()
    if role == 'a':
        admin = Admin.adminManage.get(ano=id)
        admin.password = npwd
        admin.save()

    return HttpResponse(json.dumps({
        'code': 20000,
        'data': {
            'message': 'ok'
        }
    }))

def check_password(request):
    id = eval(request.body)['id']
    # print(id)
    pwd = eval(request.body)['pwd']
    # print(pwd)
    role = eval(request.body)['role']
    # print(role)

    obj = hashlib.md5(id.encode())  # 实例化md5的时候可以给传个参数，这叫加盐
    obj.update(pwd.encode("utf-8"))  # 是再加密的时候传入自己的一块字节，
    pwd = obj.hexdigest()

    if role == 's':
        student = Student.studentManager.get(sno=id)
        if student.password == pwd:
            return HttpResponse(json.dumps({
                'code': 20000,
                'data': {
                    'message': 'ok'
                }
            }))
    if role == 't':
        teacher = Teacher.teacherManage.get(tno=id)
        if teacher.password == pwd:
            return HttpResponse(json.dumps({
                'code': 20000,
                'data': {
                    'message': 'ok'
                }
            }))
    if role == 'a':
        admin = Admin.adminManage.get(ano=id)
        if admin.password == pwd:
            return HttpResponse(json.dumps({
                'code': 20000,
                'data': {
                    'message': 'ok'
                }
            }))
    return HttpResponse(json.dumps({
        'code': 20000,
        'data': {
            'message': 'error'
        }
    }))

def updata_icon(request):
    userId = request.POST.get('userId')
    role = request.POST.get('role')
    # print(role)
    path = r'/upfile/'
    if role == 's':
        file_content = ContentFile(request.FILES['avatar'].read())
        img = StudentImage(name=request.FILES['avatar'].name, src=request.FILES['avatar'])
        img.save()
        path = r'/upfile/studentImage/' + request.FILES['avatar'].name
        stu = Student.studentManager.filter(sno=userId)
        stu = stu[0]
        stu.image = path
        stu.save()

        return HttpResponse(json.dumps({
            'code': 20000,
            'data': {
                'message': 'ok',
                'avatar': path
            }
        }))

    if role == 't':
        file_content = ContentFile(request.FILES['avatar'].read())
        img = TeacherImage(name=request.FILES['avatar'].name, src=request.FILES['avatar'])
        img.save()
        path = r'/upfile/teacherImage/' + request.FILES['avatar'].name
        tea = Teacher.teacherManage.get(tno=userId)
        tea.image = path
        tea.save()

        return HttpResponse(json.dumps({
            'code': 20000,
            'data': {
                'message': 'ok',
                'avatar': path
            }
        }))

    if role == 'a':
        file_content = ContentFile(request.FILES['avatar'].read())
        img = AdminImage(name=request.FILES['avatar'].name, src=request.FILES['avatar'])
        img.save()
        path = r'/upfile/adminImage/' + request.FILES['avatar'].name
        admin = Admin.adminManage.get(ano=userId)
        admin.image = path
        admin.save()

        return HttpResponse(json.dumps({
            'code': 20000,
            'data': {
                'message': 'ok',
                'avatar': path
            }
        }))
    return HttpResponse(json.dumps({
        'code': 20000,
        'data': {
            'message': 'error',
            'avatar': path
        }
    }))

def get_studentinfo(request):
    id = request.GET.get('id')
    student = Student.studentManager.filter(sno=id)
    student = student[0]
    return HttpResponse(json.dumps({
        'code': 20000,
        'data': {
            'stu_grade': student.grade,
            'stu_class': student.cs
        }
    }))

def create_course(request):
    id = request.POST.get('id')
    teacher = Teacher.teacherManage.get(tno=id)
    major = request.POST.get('major')
    department = Department.departmentManage.get(dno=major)
    title = request.POST.get('title')
    intro = request.POST.get('intro')

    file_content = ContentFile(request.FILES['cover'].read())
    img = ImageStore(name=request.FILES['cover'].name, img=request.FILES['cover'])
    img.save()
    imagepath = r'/upfile/course_img/' + request.FILES['cover'].name

    # cno = None
    # count = len(Course.courseManager.all()) + 1
    # cs = Course.courseManager.last().no[7]
    # print(Course.courseManager.last())
    if Course.courseManager.last() is None:
        cs = '0'
    else:
        cs = Course.courseManager.last().pk + 1
    # if cs == '0' and int(cs) < 9:
    #     cno = "2019070" + str(int(Course.courseManager.last().no[7]) + 1)
    # else:
    #     cno = "201907" + str(int(Course.courseManager.last().no[6:]) + 1)

    # if count < 10:
    #     cno = "2019070" + str(count)
    # else:
    #     cno = '201907' + str(count)
    #     no, name, tno, dno, intro, image = ''
    course = Course.courseManager.createCourse(str(cs), title, teacher, department, intro, image=imagepath)
    course.save()
    response = HttpResponse(json.dumps({
        'code': 20000,
        'data': {
            'message': 'ok'
        }
    }))
    response['Access - Control - Allow - Origin'] = "*"
    response["Access - Control - Allow - Methods"] = "POST, GET, OPTIONS"
    response["Access - Control - Max - Age"] = "1000"
    response["Access - Control - Allow - Headers"] = "*"
    response["X-Frame-Options"] = "allow-from *"
    return response

def delete_course(request):
    course_id = request.GET.get('course_id')
    try:
        course = Course.courseManager.get(no=course_id)
        delete_course_cascade(course)
    except:
        pass

    return HttpResponse(json.dumps({
        'code': 20000,
        'data': {
            'message': 'ok'
        }
    }))

def delete_course_cascade(course):
    # 删除课程对应的缩略图
    fileDir = os.path.join(settings.MEDIA_ROOT, 'course_img')
    thumbnail_path = os.path.join(fileDir, course.image.split('/')[-1])
    mclass = Mclass.mclassManager.get(isDelete=False, con=course.pk)
    try:
        os.remove(thumbnail_path)
    except Exception as e:
        print(e)

    # 删除课程对应的资源
    files = File.fileManager.filter(no=course.pk)
    for file_item in files:
        print(file_item)
        fileDir = os.path.join(settings.MEDIA_ROOT, 'source')
        filePath = os.path.join(fileDir, file_item.url.split('/')[-1])
        file_item.isDelete = True
        file_item.delete()
        try:
            os.remove(filePath)
        except Exception as e:
            print(e)
    
    # 删除学生以及删除学生的路由
    students = Student.studentManager.filter(isDelete=False, mclass=mclass)
    for student in students:
        for index in range(1,10):
            try:
                route = StudentRoutesMap.studentRoutesMapManager.get(stu_no=student.sno)
                route.delete()
            except:
                pass
        student.isDelete = True
        student.delete()
    
    # 删除class
    mclass.delete()
    
    # 删除课程下的所有任务
    try:
        homeworks = Homework.homeworkManager.filter(no=course)
        for homework in homeworks:
            try:
                work = Work.workManager.get(hno=homework)
                fileDir = os.path.join(settings.MEDIA_ROOT, 'teacher_task')
                filePath = os.path.join(fileDir, work.url.split('/')[-1])
                os.remove(filePath)
            except Exception as e:
                print(e)
                pass
            if work is not None:
                work.isDelete = True
                work.delete()
            if homework is not None:
                homework.isDelete = True
                homework.delete()
    except Exception as e:
        print(e)
        pass
    # print('hello')
    try:
        course.isDelete = True
        course.delete()
    except Exception as e:
        print(e)
        pass
    # print('hello')



def admin_getteacherinfo(request):
    major_id = request.GET.get('major_id')
    department = Department.departmentManage.get(dno=major_id)
    course_id = request.GET.get('course_id')
    page = int(request.GET.get('page'))
    limit = int(request.GET.get('limit'))
    department = Department.departmentManage.get(dno=major_id)
    teachers = Teacher.teacherManage.filter(isDelete=False, department=department)

    slen = len(teachers)
    # print(slen)
    if int(page) * int(limit) <= slen:
        teachers = teachers[(int(page) - 1) * int(limit):(int(page) - 1) * limit + limit]
        # print(teachers)
    if int(page) * int(limit) - slen < int(limit) and int(page) * int(limit) > slen:
        teachers = teachers[(int(page) - 1) * limit:slen]
        # print(teachers)
    if int(page) * int(limit) - slen > limit:
        return HttpResponse(json.dumps({
            'code': 20000,
            'data': {
                'listTotal': 0,
                'list': [],
            }
        }))
    list = []
    count = len(teachers)
    for item in teachers:
        courses = item.course_set.all()
        cslist = []
        for cs in courses:
            cslist.append(cs.name)
        dic = {
            'tea_number': item.tno,
            'tea_name': item.tname,
            'tea_position': item.botany,
            'tea_major': item.department.dname,
            'teacher_course': cslist,
        }
        list.append(dic)
    return HttpResponse(json.dumps({
        'code': 20000,
        'data': {
            'total': slen,
            'teachers': list,
        }
    }))

def admin_setpassword(request):
    # # print(json.loads(request.body))
    # admin_id = json.loads(request.body)['tea']
    tea_id = json.loads(request.body)['item']['tea_number']
    newPwd = json.loads(request.body)['item']['newPwd']
    
    obj = hashlib.md5(tea_id.encode())  # 实例化md5的时候可以给传个参数，这叫加盐
    obj.update(newPwd.encode("utf-8"))  # 是再加密的时候传入自己的一块字节，
    newPwd = obj.hexdigest()

    teacher = Teacher.teacherManage.get(tno=tea_id,isDelete=False)
    # print(newPwd)
    teacher.password = newPwd
    teacher.save()
    return HttpResponse(json.dumps({
        'code': 20000,
        'data': {
            'message': 'ok'
        }
    }))

def admin_delteacher(request):
    teacher_id = request.GET.get('id')
    teacher = Teacher.teacherManage.get(tno=teacher_id, isDelete=False)
    try:
        courses = Course.courseManager.filter(tno=teacher)
        for course in courses:
            delete_course_cascade(course)
        staging_files = StagingFile.stagingFileManager.filter(tno=teacher)
        for staging_file in staging_files:
            staging_convert_file = StagingConvertPDF.stagingConvertPDFManager.get(staging_pdf_name=staging_file.filename)
            staging_convert_file_path = os.path.join(settings.MEDIA_ROOT, staging_convert_file.staging_pdf_path)
            staging_file_path = os.path.join(settings.MEDIA_ROOT, staging_file.url)
            os.remove(staging_convert_file_path)
            os.remove(staging_file_path)
            staging_convert_file.delete()
            staging_file.delete()

    except Exception as e:
        print(e)
        

    teacher.isDelete = True
    teacher.save()

    return HttpResponse(json.dumps({
        'code':20000,
        'data': {
            'message': 'ok'
        }
    }))



#导入老师
def load_teacherinfo(request):
    major_id = request.POST.get('major_id')
    file_content = ContentFile(request.FILES['excel'].read())
    teacherStore = TeacherStore(name=request.FILES['excel'].name, src=request.FILES['excel'])
    teacherStore.save()
    fileDir = os.path.join(settings.MEDIA_ROOT, 'teacherStore')
    teacherstorepath = os.path.join(fileDir, request.FILES['excel'].name)
    # 打开文件
    file = load_workbook(filename=teacherstorepath)
    # 所有表格的名称
    sheets = file.sheetnames
    # 拿出一个表格
    sheet = file.worksheets[0]
    # 读取数据
    mclass = None
    for lineNum in range(1, sheet.max_row + 1):
        lineList = []
        for columnNum in range(1, sheet.max_column + 1):
            # 拿数据
            value = sheet.cell(row=lineNum, column=columnNum).value
            lineList.append(value)
        if lineNum > 1:
            dept = Department.departmentManage.get(dno=major_id)
            try:
                obj = hashlib.md5(lineList[0].encode())  # 实例化md5的时候可以给传个参数，这叫加盐
                obj.update("123456".encode("utf-8"))  # 是再加密的时候传入自己的一块字节，
                secret = obj.hexdigest()
                teacher = Teacher.teacherManage.createTeacher(lineList[0], lineList[1], '/upfile/template/icon.jpg', lineList[2], secret, '', dept)
                teacher.save()
                # for index in range(1, 7):
                #     route = Routes.routeManager.get(pk=index)
                #     route2 = Routes.routeManager.get(pk=index+10)

                #     teacher_map = TeacherRoutesMap.teacherRoutesMapManager.createTeacherRoutesMap(teacher, route)
                #     teacher_map2 = TeacherRoutesMap.teacherRoutesMapManager.createTeacherRoutesMap(teacher, route2)

                #     teacher_map.save()
                #     teacher_map2.save()
            except:
                pass
    return HttpResponse(json.dumps({
        'code': 20000,
        'data': {
            'message': 'ok'
        }
    }))

def search_teacher(request):
    name = request.GET.get('name')
    teachers = Teacher.teacherManage.filter(tname=name)

    list = []
    count = len(teachers)
    for item in teachers:
        courses = item.course_set.all()
        cslist = []
        for cs in courses:
            cslist.append(cs.name)
        dic = {
            'tea_number': item.tno,
            'tea_name': item.tname,
            'tea_position': item.botany,
            'tea_major': item.department.dname,
            'teacher_course': cslist,
        }
        list.append(dic)
    return HttpResponse(json.dumps({
        'code': 20000,
        'data': {
            'total': count,
            'teachers': list,
        }
    }))

def search_teacher_with_no(request):
    tea_no = request.GET.get('teaNo')
    teachers = Teacher.teacherManage.filter(tno=tea_no)

    array = []
    count = len(teachers)
    for item in teachers:
        courses = item.course_set.all()
        courseList = []
        for course in courses:
            courseList.append(course.name)
        dic = {
            'tea_number': item.tno,
            'tea_name': item.tname,
            'tea_position': item.botany,
            'tea_major': item.department.dname,
            'teacher_course': courseList
        }
        array.append(dic)

    return HttpResponse(json.dumps({
        'code': 20000,
        'data': {
            'total': count,
            'teachers': array
        }
    }))


def get_teachertemplate(request):
    return HttpResponse(json.dumps({
        'code': 20000,
        'data': {
            'link': r'/upfile/template/teachers.xlsx'
        }
    }))

def search_course(request):
    course_name = request.GET.get("course_name")
    # print(major_id)
    # page = request.GET.get('page')
    # limit = request.GET.get('limit')
    courses = Course.courseManager.filter(name__contains=course_name)
    # department = Department.departmentManage.get(dno=major_id)
    # courses = department.course_set.all()
    # clen = len(courses)
    # if int(page) * int(limit) < clen:
    #     courses = courses[(int(page) - 1) * 5:(int(page) - 1) * 5 + 5]
    # if int(page) * int(limit) - clen < 5 and int(page) * int(limit) > clen:
    #     courses = courses[(int(page) - 1) * 5:clen - (int(page) - 2) * 5]
    # if int(page) * int(limit) - clen > 5:
    #     return HttpResponse(json.dumps({
    #         'code': 20000,
    #         'data': {
    #             'total': 0,
    #             'courses': [],
    #         }
    #     }))

    list = []
    count = len(courses)
    for item in courses:
        course_count = len(item.file_set.all())
        dic = {
            'course_cover': item.image,
            'course_name': item.name,
            'course_teacher': item.tno.tname,
            'course_id': item.no,
            'upload_count': course_count,
        }
        list.append(dic)
    return HttpResponse(json.dumps({
        'code': 20000,
        'data': {
            'total': count,
            'courses': list,
        }
    }))

def delete_student(request):
    id = eval(request.body)['id']
    course = eval(request.body)['course']
    course_id = Course.courseManager.get(no=course)
    mclass = Mclass.mclassManager.get(con=course_id)
    student = Student.studentManager.get(sno=id, mclass=mclass)
    # student.isDelete = True
    # student.save()
    student.delete()
    return HttpResponse(json.dumps({
        'code': 20000,
        'data': {
            'message': "ok",
        }
    }))

from .models import Work, Homework
#发布任务
def publish_task(request):
    describe = request.POST.get('describe')
    type = request.POST.get('type')
    source_type = Source.sourceManager.filter(sno=type)[0]
    title = request.POST.get('title')
    course_id = request.POST.get('course_id')
    course = Course.courseManager.filter(no=course_id)[0]
    count = request.POST.get('count')

    home_work = Homework.homeworkManager.createHomework(describe, title, source_type,course)
    home_work.save()
    # fileList = []
    for i in range(int(count)):
        file1 = request.FILES.get('file' + str(i))
        fileDir = os.path.join(settings.MEDIA_ROOT, 'teacher_task')
        filePath = os.path.join(fileDir, file1.name)
        with open(filePath, 'wb') as fp:
            for info in file1.chunks():
                fp.write(info)
        # print(os.path.join('/upfile/source', file1.name))
        work = Work.workManager.createWork(file1.name, r'/upfile/teacher_task/' + file1.name, home_work)
        work.save()
        # fileList.append(f)
    return HttpResponse(json.dumps({
        'code': 20000,
        'data': {
            'message': 'ok'
        }
    }))

#请求发布的任务
def get_homework(request):
    course_id = request.GET.get('course_id')
    course = Course.courseManager.filter(no=course_id)[0]
    page = int(request.GET.get('page'))
    limit = int(request.GET.get('limit'))

    homeworks = course.homework_set.all()
    slen = len(homeworks)
    # print(slen)
    if int(page) * int(limit) <= slen:
        homeworks = homeworks[(int(page) - 1) * int(limit):(int(page) - 1) * limit + limit]
    if int(page) * int(limit) - slen < int(limit) and int(page) * int(limit) > slen:
        homeworks = homeworks[(int(page) - 1) * limit:slen]
    if int(page) * int(limit) - slen > limit:
        return HttpResponse(json.dumps({
            'code': 20000,
            'data': {
                'listTotal': 0,
                'list': [],
            }
        }))
    list = []

    for item in homeworks:
        dic = {
            'task_id': item.pk,
            'release_time': str(item.time).split()[0],
            'task_title': item.title,
            'task_intro': item.describe,
            'task_type': item.sno.sname,
        }
        list.append(dic)

    return HttpResponse(json.dumps({
        'code': 20000,
        'data': {
            'listTotal': slen,
            'list': list
        }
    }))

#请求任务下的附件
def get_work(request):
    task_id = request.GET.get('task_id')

    task = Homework.homeworkManager.get(pk=int(task_id))
    works = task.work_set.all()

    fileList = []
    for item in works:
        dic = {
            'task_title': task.title,
            'file_id': item.pk,
            'task_filelink': item.url,
            'task_filename': item.title
        }
        fileList.append(dic)
    return HttpResponse(json.dumps({
        'code': 20000,
        'data': {
            'fileList': fileList
        }
    }))

def teacherget_course(request):
    id = request.GET.get('id')
    major_id = request.GET.get('major')
    teacher = Teacher.teacherManage.get(tno=id)
    department = Department.departmentManage.get(dno=major_id)
    course = teacher.course_set.all()
    courses = []
    for item in course:
        if item.dno == department:
            dic = {
                'title': item.name,
                'course_id': item.no
            }
            courses.append(dic)
        
    return HttpResponse(json.dumps({
        'code': 20000,
        'data': {
            'courses': courses,

        }
    }))

def studentget_course(request):
    id = request.GET.get('id')
    student = Student.studentManager.filter(sno=id)
    # print(student)
    mclass = []
    for item1 in student:
        mclass.append(item1.mclass)
    courses = []
    for item in mclass:
        dic = {
            'label': item.con.name,
            'value': item.con.no
        }
        courses.append(dic)
    return HttpResponse(json.dumps({
        'code': 20000,
        'data': {
            'courses': courses,

        }
    }))


def studentget_garden(request):
    id = request.GET.get('id')
    student = Student.studentManager.filter(sno=id)[0]

    grades = {
        'label': student.grade,
        'value': student.grade
    }
    majors = {
        'label': student.department.dname,
        'value': student.department.dno
    }
    return HttpResponse(json.dumps({
        'code': 20000,
        'data': {
            'grades': [grades],
            'majors': [majors],
        }
    }))

def teacherget_garden(request):
    id = request.GET.get('id')
    teacher = Teacher.teacherManage.get(tno=id)
    course = teacher.course_set.all()
    grades = []
    majors = []
    itemset = set()
    for item in course:
        myclass = item.mclass_set.all()
        # print(myclass)
        if len(list(myclass)) > 0:
            myclass = myclass[0]
            str = myclass.grade + '|' +  item.dno.dname + '|' + item.dno.dno
            itemset.add(str)

    for item2 in itemset:
        dic1 = {
            'label': item2.split('|')[0],
            'value': item2.split('|')[0]
        }
        dic2 = {
            'label': item2.split('|')[1],
            'value': item2.split('|')[2]
        }
        grades.append(dic1)
        majors.append(dic2)
    return HttpResponse(json.dumps({
        'code': 20000,
        'data': {
            'grades': grades,
            'majors': majors,
        }
    }))

def gettaskbycourse(request):
    course_id = request.GET.get('course_id')
    course = Course.courseManager.get(no=course_id)
    homeworks = list(course.homework_set.all())
    tasks = []
    if len(homeworks) > 0:
        for item in homeworks:
            dic = {
                'label': item.title,
                'value': str(item.pk)
            }
            tasks.append(dic)
    return HttpResponse(json.dumps({
        'code': 20000,
        'data': {
            'tasks': tasks,
            'status': [
                {
                    'label': '全部',
                    'value': 'all'
                },
                {
                    'label': '已批改',
                    'value': 'success'
                },
                {
                    'label': '未批改',
                    'value': 'info'
                }
            ]
        }
    }))

#提交记录表格数据
def getTableData(request):
    sno = request.GET.get('id')
    course = request.GET.get('course')
    # 查出当前stu_id下的提交记录
    submit_records = StudentWork.studentWorkManager.filter(stu_id=sno)
    records = []
    for item in submit_records:
        work_it = item.work
        course_it = work_it.no
        dic = {
            'stu_submitDate': str(item.stu_submitDate),
            'stu_taskTitle': work_it.title,
            'stu_course': course_it.name
        }
        records.append(dic)
    if len(course)>0:
        course_name = Course.courseManager.get(no=course).name
        for item in records:
            if item.get('stu_course') != course_name:
                records.remove(item)
    return HttpResponse(json.dumps({
        'code': 20000,
        'data': {
            'total': len(records),
            'message': 'ok',
            'list': records
        }
    }))

#删除任务
def del_homework(request):
    task_id = int(request.GET.get('task_id'))
    homework = Homework.homeworkManager.get(pk=task_id)

    works = homework.work_set.all()
    for item in works:
        fileDir = os.path.join(settings.MEDIA_ROOT, 'teacher_task')
        studentstorepath = os.path.join(fileDir, item.url.split('/')[-1])
        os.remove(studentstorepath)
        item.delete()

    homework.delete()

    return HttpResponse(json.dumps({
        'code': 20000,
        'data': {
            'message': 'ok'
        }
    }))

#删除任务附件
def del_work(request):
    task_id = int(request.GET.get('task_id'))
    file_id = int(request.GET.get('file_id'))

    work = Work.workManager.get(pk=int(file_id))
    fileDir = os.path.join(settings.MEDIA_ROOT, 'teacher_task')
    studentstorepath = os.path.join(fileDir, work.url.split('/')[-1])
    os.remove(studentstorepath)
    work.delete()

    return HttpResponse(json.dumps({
        'code': 20000,
        'data': {
            'message': 'ok'
        }
    }))



#数据模拟
from faker import Faker
import random
# createTeacher(self, tno, tname, botany, password, email, department, isDelete=False)
def teacher_data(request):
    f = Faker('zh_CN')
    botanys = ['讲师', '副教授', '教授']
    for i in range(20):
        temp = i + 2
        if temp > 9:
            tno = "22203200" + str(temp)
        else:
            tno = "222032000" + str(temp)
        tname = f.name()
        # print(random.randint(0, 3))
        botany = botanys[random.randint(0, 2)]
        password = '123456'
        email = '908765543@qq.com'
        # print((i % 4))
        department = Department.departmentManage.get(pk=((i % 4) + 1))
        t = Teacher.teacherManage.createTeacher(tno, tname, botany, password, email, department)
        t.save()
    return HttpResponse('successful')

def update_stus_teacher(request):
    students = Student.studentManager.all()
    for item in students:
        # print(newPwd)
        obj = hashlib.md5(item.sno.encode())  # 实例化md5的时候可以给传个参数，这叫加盐
        obj.update(item.password.encode("utf-8"))  # 是再加密的时候传入自己的一块字节，
        newPwd = obj.hexdigest()
        item.password = newPwd
        item.save()
    teachers = Teacher.teacherManage.all()
    for it in teachers:
        # print(newPwd)
        obj = hashlib.md5(it.tno.encode())  # 实例化md5的时候可以给传个参数，这叫加盐
        obj.update(it.password.encode("utf-8"))  # 是再加密的时候传入自己的一块字节，
        newPwd = obj.hexdigest()
        it.password = newPwd
        it.save()
    return HttpResponse('successful')

from django.core.files.base import ContentFile
def getImg(request):
    file_content = ContentFile(request.FILES['file'].read())
    img = ImageStore(name = request.FILES['file'].name, img =request.FILES['file'])
    img.save()
    return HttpResponse('successful')